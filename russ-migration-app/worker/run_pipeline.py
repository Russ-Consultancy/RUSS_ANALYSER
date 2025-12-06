import argparse
import pandas as pd
import re
from pathlib import Path
import shutil
import subprocess
import openpyxl
from openpyxl import load_workbook

# ---------------------------------------------------------------------
BASE = Path(__file__).resolve().parent
UPLOADS = BASE.parent / "uploads"
OUTPUTS = BASE.parent / "outputs"
TEMPLATE_XLSX = BASE.parent / "analysis_templates" / "analysis_template.xlsx"

OUTPUTS.mkdir(exist_ok=True)

# ---------------------------------------------------------------------
def extract_db_name(html_text):
    for pat in [
        r"Database Name\s*[:=]\s*([A-Za-z0-9_]+)",
        r"DB Name\s*[:=]\s*([A-Za-z0-9_]+)",
    ]:
        m = re.search(pat, html_text, re.IGNORECASE)
        if m:
            return m.group(1).strip()
    return "UNKNOWN_DB"

# ---------------------------------------------------------------------
def regex_fallback_parser(html_file: Path):
    """Improved fallback parser for single AWR HTML reports."""
    text = html_file.read_text(errors="ignore")

    def find(pat, multiplier=1):
        m = re.search(pat, text, re.IGNORECASE)
        return float(m.group(1).replace(",", "")) * multiplier if m else 0.0

    sga_mb = find(r"SGA use\s*\(MB\)\s*[:=]?\s*([\d,\.]+)")
    if not sga_mb:
        sga_mb = find(r"SGA Target\s*\(GB\)\s*[:=]?\s*([\d,\.]+)", 1024)

    pga_mb = find(r"PGA use\s*\(MB\)\s*[:=]?\s*([\d,\.]+)")
    if not pga_mb:
        pga_mb = find(r"PGA Aggregate Target\s*\(GB\)\s*[:=]?\s*([\d,\.]+)", 1024)

    return {
        "SGA_MB": sga_mb,
        "PGA_MB": pga_mb,
        "PHYS_READ_REQ": find(r"Physical read total IO requests\s*[:=]?\s*([\d,\.]+)"),
        "PHYS_WRITE_REQ": find(r"Physical write total IO requests\s*[:=]?\s*([\d,\.]+)"),
        "PHYS_READ_MB": find(r"Physical read total bytes\s*[:=]?\s*([\d,\.]+)", 1 / (1024 * 1024)),
        "PHYS_WRITE_MB": find(r"Physical write total bytes\s*[:=]?\s*([\d,\.]+)", 1 / (1024 * 1024)),
        "CPU": find(r"CPU Count\s*[:=]?\s*([\d,\.]+)"),
        "MEMORY_GB": find(r"Memory\s*\(GB\)\s*[:=]?\s*([\d,\.]+)"),
        "ELAPSED": find(r"Elapsed\s*\(min\)\s*[:=]?\s*([\d,\.]+)"),
        "DBTIME": find(r"DB Time\s*\(min\)\s*[:=]?\s*([\d,\.]+)"),
    }

# ---------------------------------------------------------------------
def extract_metrics_from_output_xlsx(output_xlsx: Path):
    """Read metrics directly from the generated output Excel."""
    wb = openpyxl.load_workbook(output_xlsx, data_only=True)
    ws = wb.active

    def val(cell):
        return ws[cell].value if ws[cell].value is not None else 0

    return {
        "DB_NAME": val("N2"),
        "ELAPSED": val("AS2"),
        "DBTIME": val("AT2"),
        "CPU": val("S2"),
        "CORES": val("S2") / 2 if val("S2") else 0,
        "MEMORY_GB": val("W2"),
        "SGA_MB": val("X2"),
        "PGA_MB": val("Y2"),
        "PHYS_READ_MB": val("AE2"),
        "PHYS_WRITE_MB": val("AF2"),
        "PHYS_READ_REQ": val("AB2"),
        "PHYS_WRITE_REQ": val("AC2"),
    }

# ---------------------------------------------------------------------
def write_to_template_from_output(metrics):
    db_name = metrics["DB_NAME"] or "UNKNOWN_DB"
    out_file = OUTPUTS / f"final_analysis_{db_name}.xlsx"
    shutil.copy(TEMPLATE_XLSX, out_file)
    wb = load_workbook(out_file)

    ws_data = wb["AWRData"]

    ws_data["A2"] = metrics["DB_NAME"]
    ws_data["D2"] = metrics["ELAPSED"]
    ws_data["E2"] = metrics["DBTIME"]
    ws_data["F2"] = metrics["CPU"]
    ws_data["G2"] = metrics["CORES"]
    ws_data["H2"] = metrics["MEMORY_GB"]
    ws_data["I2"] = metrics["SGA_MB"]
    ws_data["J2"] = metrics["PGA_MB"]
    ws_data["K2"] = metrics["PHYS_READ_MB"]
    ws_data["L2"] = metrics["PHYS_WRITE_MB"]
    ws_data["M2"] = metrics["PHYS_READ_REQ"]
    ws_data["N2"] = metrics["PHYS_WRITE_REQ"]

    # Database Analysis sheet
    if "Database Analysis" in wb.sheetnames:
        ws_analysis = wb["Database Analysis"]
        ws_analysis["A2"] = metrics["DB_NAME"]

    wb.save(out_file)
    print(f"✅ final_analysis.xlsx written for {db_name}")
    return out_file

# ---------------------------------------------------------------------
def run(awr_file):
    awr_path = Path(awr_file)
    print(f"⚙️ Processing {awr_path.name} ...")

    # 1️⃣ Create temp input/output folders for process_awr_reports.py
    temp_input_dir = OUTPUTS / f"tmp_{awr_path.stem}"
    temp_input_dir.mkdir(exist_ok=True)
    temp_output_dir = OUTPUTS / f"{awr_path.stem}"
    temp_output_dir.mkdir(exist_ok=True)

    # Copy HTML file into the input directory
    shutil.copy(awr_path, temp_input_dir / awr_path.name)

    print(f"🧩 Launching process_awr_reports.py on {awr_path.name} ...")
    subprocess.run(
        ["python3", str(BASE / "process_awr_reports.py"), str(temp_input_dir), str(temp_output_dir)],
        check=True
    )

    # 2️⃣ Detect the output Excel
    xlsx_files = list(temp_output_dir.glob("*.xlsx"))
    if not xlsx_files:
        raise FileNotFoundError(f"No Excel output found in {temp_output_dir}")

    output_xlsx = xlsx_files[0]
    print(f"✅ Found AWR output: {output_xlsx.name}")

    # 3️⃣ Extract and write to template
    metrics = extract_metrics_from_output_xlsx(output_xlsx)

    print("🧾 Extracted Metrics:")
    for k, v in metrics.items():
        print(f"   {k}: {v}")

    return write_to_template_from_output(metrics)

# ---------------------------------------------------------------------
if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--awr-file", required=True)
    args = parser.parse_args()
    run(args.awr_file)
