#!/usr/bin/env python3
from pathlib import Path
from openpyxl import load_workbook
import csv
import datetime

BASE = Path(__file__).resolve().parent
OUTPUTS = BASE.parent / "outputs"
OUTPUTS.mkdir(exist_ok=True)

xlsx_files = sorted(list(OUTPUTS.glob("final_analysis_*.xlsx")))

if not xlsx_files:
    print("No final_analysis_*.xlsx files found in outputs/. Run the pipeline first.")
    raise SystemExit(1)

debug_rows = []
ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
csv_path = OUTPUTS / f"debug_cells_{ts}.csv"

for f in xlsx_files:
    print(f"\n--- File: {f.name} ---")
    try:
        wb = load_workbook(f, data_only=False)   # data_only=False so we can see formulas too
    except Exception as e:
        print("  ERROR opening workbook:", e)
        continue

    # list candidate sheets
    sheets = wb.sheetnames
    print("  Sheets:", sheets)

    # Candidate sheet names to inspect
    candidates = []
    if "Database Analysis" in sheets:
        candidates.append("Database Analysis")
    if "AWRData" in sheets:
        candidates.append("AWRData")
    # also include the first sheet as fallback
    candidates.append(wb.sheetnames[0])

    # Candidate cells (from our conversation)
    # Database Analysis: A2..E2
    # AWRData: A2, B2, C2, D2, E2, F2, G2, H2, I2, J2, K2, L2, M2, N2, etc (we'll inspect a range)
    db_cells = ["A2","B2","C2","D2","E2"]
    awr_cells = ["A2","B2","C2","D2","E2","F2","G2","H2","I2","J2","K2","L2","M2","N2","AS2","AT2"]  # AS/AT because earlier messages referenced AS2/AT2

    for sheet_name in candidates:
        ws = wb[sheet_name]
        for cell in sorted(set(db_cells + awr_cells)):
            val = None
            formula = None
            try:
                cell_obj = ws[cell]
                val = cell_obj.value
                # openpyxl stores formula in .value when data_only=False and if the cell contains a formula
                # but sometimes .value is the formula string; we try to read both raw and data-only if available
            except Exception:
                val = "<ERR>"

            # Also try data_only workbook to get last calculated value
            try:
                wb_data = load_workbook(f, data_only=True)
                ws_data = wb_data[sheet_name] if sheet_name in wb_data.sheetnames else None
                data_val = ws_data[cell].value if ws_data is not None else None
            except Exception:
                data_val = None

            print(f"  {sheet_name} {cell} -> raw: {repr(val)}  | data_only: {repr(data_val)}")
            debug_rows.append({
                "file": f.name,
                "sheet": sheet_name,
                "cell": cell,
                "raw_value": val,
                "data_value": data_val
            })

# write CSV
with csv_path.open("w", newline="", encoding="utf-8") as fh:
    writer = csv.DictWriter(fh, fieldnames=["file","sheet","cell","raw_value","data_value"])
    writer.writeheader()
    for r in debug_rows:
        # stringify values to avoid binary issues
        writer.writerow({
            "file": r["file"],
            "sheet": r["sheet"],
            "cell": r["cell"],
            "raw_value": "" if r["raw_value"] is None else str(r["raw_value"]),
            "data_value": "" if r["data_value"] is None else str(r["data_value"]),
        })

print("\nWrote debug CSV to:", csv_path)
print("Open it and check the raw_value vs data_value for Database Analysis B2..E2 and AWRData equivalents.")
